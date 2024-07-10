# Diego Insuasty Mora (Obtiene la medida tanto horizontal como vertical de los archivos PDF antes de convertir a formato Carta)

import openpyxl
import PyPDF4
import decimal


ruta_archivo = r'\\FJCALDAS\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\tamanoarchivos24_04.xlsx'

# Abre el archivo de Excel
wb = openpyxl.load_workbook(ruta_archivo)

# Selecciona la hoja de Excel que contiene los nombres y ubicaciones de los archivos PDF
sheet = wb['Hoja1']

# Inicializa un contador para llevar un seguimiento de las filas
row_counter = 2

# Recorre las filas de la hoja de Excel a partir de la fila 2 (A2)
for row in sheet.iter_rows(min_row=2, values_only=True):

    # Obtiene el nombre y la ubicación del archivo PDF de la columna A
    file_path = row[0]

    # Abre el archivo PDF en modo de lectura binaria
    with open(file_path, 'rb') as pdf_file:

        # Crea un objeto de lectura de PDF
        pdf_reader = PyPDF4.PdfFileReader(pdf_file)

        # Obtiene el tamaño de la primera página del PDF
        page_size = pdf_reader.getPage(0).mediaBox

        # Determina el tamaño de página
        page_width = round(page_size.getWidth() / 72, 2) # Convierte de puntos a pulgadas y redondea a dos decimales
        page_height = round(page_size.getHeight() / 72, 2) # Convierte de puntos a pulgadas y redondea a dos decimales

        # Escribe el tamaño de la página en centímetros en las columnas B y C
        sheet.cell(row=row_counter, column=2).value = round(decimal.Decimal(str(page_height)) * decimal.Decimal('2.54'), 2)
        sheet.cell(row=row_counter, column=3).value = round(decimal.Decimal(str(page_width)) * decimal.Decimal('2.54'), 2)
        
        if page_width == 8.5 and page_height == 11:
            sheet.cell(row=row_counter, column=4).value = 'Carta'
        elif page_width == 8.5 and page_height == 14:
            sheet.cell(row=row_counter, column=4).value = 'Oficio'
        else:
            sheet.cell(row=row_counter, column=4).value = 'Otro tamaño'

        # Incrementa el contador de filas
        row_counter += 1

# Guarda los cambios en el archivo de Excel
wb.save(r'\\FJCALDAS\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\tamanoarchivos24_04.xlsx')
print('Archivo terminado.')





