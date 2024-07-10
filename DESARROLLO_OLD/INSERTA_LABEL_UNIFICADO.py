# v.2.2 AUTOR JOHN GUTIERREZ ("***INSERTA STIKER lABEL MULTIPLES FORMATOS DE PDF***")
import os
import io
from openpyxl import load_workbook
from PyPDF4 import PdfFileReader, PdfFileWriter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm

# cargar el archivo de Excel
workbook = load_workbook(r'\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\ARCHIVOS_FUENTE.xlsx')   
worksheet = workbook.active

# obtener la última fila de datos
max_row = worksheet.max_row

# iterar a través de las filas de datos
for row in range(1, max_row + 1):
    # obtener los datos del archivo de Excel
    pdf_name = worksheet.cell(row, 1).value
    pdf_path = worksheet.cell(row, 2).value
    sticker_name = worksheet.cell(row, 3).value
    sticker_path = worksheet.cell(row, 4).value
    
    # abrir el archivo PDF y crear un nuevo archivo PDF actualizado
    with open(pdf_path, 'rb') as pdf_file:
        pdf_reader = PdfFileReader(pdf_file)
        pdf_writer = PdfFileWriter()
        
        # agregar la pegatina solo en la primera página
        is_first_page = True
        
        # iterar a través de las páginas del archivo PDF
        for page_num in range(pdf_reader.getNumPages()):
            page = pdf_reader.getPage(page_num)
            
            # agregar la pegatina solo en la primera página
            if is_first_page:
                # agregar la pegatina a la página actual
                packet = io.BytesIO()
                can = canvas.Canvas(packet, pagesize=letter)

                # Obtener la altura de la página actual en centímetros
                page_height_cm = float(page.mediaBox.getHeight()) / 28.35

                if page_height_cm > 42:
                    can.drawImage(sticker_path, 6, 40.50*cm, 7.2*cm, 2.4*cm)                
                elif page_height_cm > 39:
                    can.drawImage(sticker_path, 6, 39.10*cm, 7.2*cm, 2.4*cm)
                elif page_height_cm > 35:
                    can.drawImage(sticker_path, 6, 32.95*cm, 7.2*cm, 2.4*cm)
                elif page_height_cm > 33:
                    can.drawImage(sticker_path, 6, 30.35*cm, 7.2*cm, 2.4*cm)
                elif page_height_cm > 29:
                    can.drawImage(sticker_path, 6, 27.15*cm, 7.2*cm, 2.4*cm)
                elif page_height_cm > 27:
                    can.drawImage(sticker_path, 6, 25.35*cm, 7.2*cm, 2.4*cm)
                else:
                    can.drawImage(sticker_path, 6, 23.75*cm, 7.2*cm, 2.4*cm)               
                
                can.save()

                # mover al comienzo del paquete y agregarlo a la página actual
                packet.seek(0)
                overlay = PdfFileReader(packet).getPage(0)
                page.mergePage(overlay)
                
                # actualizar la variable is_first_page para que las páginas posteriores no tengan pegatinas
                is_first_page = False
            
            # agregar la página actualizada al archivo PDF actualizado
            pdf_writer.addPage(page)
        
        # guardar el archivo PDF actualizado
        output_filename = os.path.join(r'\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\PRUEBA11032024\NUEVOS ORIGINAL', pdf_name)
        with open(output_filename, 'wb') as output:
            pdf_writer.write(output)
print("el archivo ha sido escrito satisfactoriamente")