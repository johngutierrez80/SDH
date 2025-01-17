import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import fitz  # Importar PyMuPDF
from math import ceil  # Importar la función ceil para redondear hacia arriba

ubicacion_carpeta = r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\Actos GG_ Diarios_SHD_2024\240729-DIB+DCO-GG-lote 1\RENOMBRADOS"

if not os.path.exists(ubicacion_carpeta):
    print(f"El directorio {ubicacion_carpeta} no existe.")
else:
    archivos_pdf = [archivo for archivo in os.listdir(ubicacion_carpeta) if archivo.endswith(".pdf")]

# Cargar el archivo Excel existente
archivo_excel = r'\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\14_ALEJANDRO\NUM_PAG.xlsx'
wb = load_workbook(archivo_excel)
ws = wb.active

# Obtener la columna correspondiente a la letra
def get_column_index(column_letter):
    return ord(column_letter.upper()) - 64

# Obtener la letra correspondiente a la columna
def get_column_letter(column_index):
    return chr(column_index + 64)

# Obtener la primera columna vacía para escribir los datos
def get_next_empty_row(column_letter):
    column_index = get_column_index(column_letter)
    for row in range(2, ws.max_row + 1):
        cell_value = ws[get_column_letter(column_index) + str(row)].value
        if cell_value is None:
            return row
    return ws.max_row + 1

total_paginas = 0
total_redondeo = 0

# Recorrer los archivos PDF y obtener el número de páginas
for archivo_pdf in archivos_pdf:
    archivo_completo = os.path.join(ubicacion_carpeta, archivo_pdf)
    nombre_archivo = os.path.basename(archivo_pdf)

    try:
        pdf = fitz.open(archivo_completo)  # Abrir el archivo con PyMuPDF
        numero_paginas = pdf.page_count
        redondeo_maximo = ceil(numero_paginas / 2)  # Redondear hacia arriba
        total_paginas += numero_paginas
        total_redondeo += redondeo_maximo
    except:
        numero_paginas = 'Error al obtener el número de páginas'
        redondeo_maximo = 'Error en el cálculo del redondeo máximo'
    
    # Obtener la próxima fila vacía para escribir los datos
    siguiente_fila = get_next_empty_row('A')
    
    # Escribir los datos en el archivo Excel
    ws['A' + str(siguiente_fila)] = nombre_archivo
    ws['B' + str(siguiente_fila)] = numero_paginas
    ws['C' + str(siguiente_fila)] = redondeo_maximo

# Obtener la próxima fila vacía para escribir los totales
siguiente_fila = get_next_empty_row('A')

# Escribir los totales en el archivo Excel
ws['A' + str(siguiente_fila)] = "Total"
ws['B' + str(siguiente_fila)] = total_paginas
ws['C' + str(siguiente_fila)] = total_redondeo

# Multiplicar los totales por 2 y escribirlos en las siguientes celdas
ws['B' + str(siguiente_fila + 1)] = total_paginas * 2
ws['C' + str(siguiente_fila + 1)] = total_redondeo * 2

# Escribir la etiqueta "TOTAL FINALES" en la columna A
ws['A' + str(siguiente_fila + 1)] = "TOTAL FINALES CLICKS/HOJAS"

# Resaltar los últimos dos valores con otro color de texto (en este caso, rojo)
red_font = Font(color="FF0000")
ws['B' + str(siguiente_fila + 1)].font = red_font
ws['C' + str(siguiente_fila + 1)].font = red_font

# Guardar el archivo Excel actualizado
wb.save(archivo_excel)
print("El archivo Excel ha sido actualizado exitosamente.")
