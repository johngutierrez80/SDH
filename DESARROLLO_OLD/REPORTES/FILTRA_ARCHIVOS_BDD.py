import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Ruta del archivo Excel a cargar en el DataFrame
excel_file = r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\LABEL_ACTOS_GG_2024\24-04-02\Extra\02_04_2024_base de datos extra.xlsx"

# Leer el archivo Excel y convertirlo en un DataFrame
df = pd.read_excel(excel_file)

# Filtrar por el tipo de "Código de dependencia" para COBROS
df_cobros = df[df['Código de dependencia'].str.contains('COBROS', case=False)]

# Filtrar por el tipo de "Código de dependencia" para IMPUESTOS
df_impuestos = df[df['Código de dependencia'].str.contains('IMPUESTOS', case=False)]

# Mostrar los DataFrames por consola
print("DataFrame para COBROS:")
print(df_cobros)

print("\nDataFrame para IMPUESTOS:")
print(df_impuestos)

# Ruta del directorio de salida
output_directory = r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\REPORTES_ACTOS_DIARIOS_SDH_2024"

# Crear el directorio de salida si no existe
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

# Rutas de salida para los archivos Excel filtrados
output_file_cobros = os.path.join(output_directory, "CANTIDADES_ACTOS_GG_COBROS.xlsx")
output_file_impuestos = os.path.join(output_directory, "CANTIDADES_ACTOS_GG_IMPUESTOS.xlsx")

# Crear un nuevo libro de trabajo de Excel
wb_cobros = Workbook()
wb_impuestos = Workbook()

# Eliminar la hoja de trabajo por defecto
default_sheet_cobros = wb_cobros.active
wb_cobros.remove(default_sheet_cobros)
default_sheet_impuestos = wb_impuestos.active
wb_impuestos.remove(default_sheet_impuestos)

# Crear hoja de trabajo para COBROS
ws_cobros = wb_cobros.create_sheet("CANTIDADES_ACTOS_GG_COBROS")
for r in dataframe_to_rows(df_cobros, index=False, header=True):
    ws_cobros.append(r)

# Crear hoja de trabajo para IMPUESTOS
ws_impuestos = wb_impuestos.create_sheet("CANTIDADES_ACTOS_GG_IMPUESTOS")
for r in dataframe_to_rows(df_impuestos, index=False, header=True):
    ws_impuestos.append(r)

# Guardar los libros de trabajo en los archivos de salida
wb_cobros.save(output_file_cobros)
wb_impuestos.save(output_file_impuestos)
print("El archivo Excel ha sido actualizado exitosamente.")