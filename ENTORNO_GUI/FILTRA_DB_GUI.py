from PIL import Image, ImageTk
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
import subprocess
import time

class MainWindow(tk.Tk):
    
    def __init__(self):
        tk.Tk.__init__(self)
        self.title("Filtrar Base de Datos")
        self.geometry("730x360")  # Establecer el tamaño de la ventana primero
        self.create_widgets()  
        self.center_window()  
        self.fade_in()  # Agregamos el efecto de fade-in al iniciar la ventana

    def create_widgets(self):
        icon_image = Image.open(r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\ENTORNO_GUI\Background\INC.ico")
        icon_photo = ImageTk.PhotoImage(icon_image)
        self.iconphoto(False, icon_photo)  

        self.background_image = Image.open(r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\ENTORNO_GUI\Background\Reportes.png")  
        self.background_photo = ImageTk.PhotoImage(self.background_image)
        self.background_label = tk.Label(self, image=self.background_photo)
        self.background_label.place(x=0, y=0, relwidth=1, relheight=1)

        self.path_label = tk.Label(self, text="Ruta de la Base de Datos:")
        self.path_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")  

        self.path_entry = tk.Entry(self, width=50)
        self.path_entry.grid(row=0, column=1, padx=10, pady=10)  

        self.select_button = tk.Button(self, text="Seleccionar BD", command=self.select_database, bg="Gray", font=("Arial", 11),width=20)
        self.select_button.grid(row=0, column=2, padx=10, pady=10)  

        self.clean_button = tk.Button(self, text="Limpiar Campos", command=self.clean_fields, bg="orange", font=("Arial", 11),width=20)
        self.clean_button.place(relx=0.2, rely=0.9, anchor="center")

        self.close_button = tk.Button(self, text="Cerrar", command=self.fade_out_and_quit, bg="red", font=("Arial", 11),width=20)
        self.close_button.place(relx=0.8, rely=0.9, anchor="center")

        self.process_button = tk.Button(self, text="Procesar", fg="white", command=self.process_data,bg="green", font=("Arial", 11),width=20)
        self.process_button.place(relx=0.5, rely=0.9, anchor="center")

        self.open_dir_button = tk.Button(self, text="Abrir Directorio", command=self.abrir_directorio, bg="gray", font=("Arial", 11),width=20)
        self.open_dir_button.place( relx=0.8, rely=0.7, anchor="center")

    def fade_in(self):
        # Efecto de fade-in gradual
        for alpha in range(0, 100):
            self.attributes('-alpha', alpha / 100)
            self.update_idletasks()
            time.sleep(0.01)

    def fade_out_and_quit(self):
        # Efecto de fade-out gradual antes de cerrar la ventana
        for alpha in range(100, 0, -1):
            self.attributes('-alpha', alpha / 100)
            self.update_idletasks()
            time.sleep(0.01)
        self.quit()

    def select_database(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.path_entry.delete(0, tk.END)
            self.path_entry.insert(0, file_path)

    def clean_fields(self):
        self.path_entry.delete(0, tk.END)

    def process_data(self):
        file_path = self.path_entry.get()
        if not file_path:
            messagebox.showerror("Error", "Por favor selecciona una base de datos.")
            return

        try:
            df = pd.read_excel(file_path)
            df_cobros = df[df['Código de dependencia'].str.contains('COBROS', case=False)]
            df_impuestos = df[df['Código de dependencia'].str.contains('IMPUESTOS', case=False)]

            output_directory = r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\REPORTES_ACTOS_DIARIOS_SDH_2024"
            if not os.path.exists(output_directory):
                os.makedirs(output_directory)

            output_file_cobros = os.path.join(output_directory, "CANTIDADES_ACTOS_GG_COBROS.xlsx")
            output_file_impuestos = os.path.join(output_directory, "CANTIDADES_ACTOS_GG_IMPUESTOS.xlsx")

            wb_cobros = Workbook()
            wb_impuestos = Workbook()

            default_sheet_cobros = wb_cobros.active
            wb_cobros.remove(default_sheet_cobros)
            default_sheet_impuestos = wb_impuestos.active
            wb_impuestos.remove(default_sheet_impuestos)

            ws_cobros = wb_cobros.create_sheet("CANTIDADES_ACTOS_GG_COBROS")
            for r in dataframe_to_rows(df_cobros, index=False, header=True):
                ws_cobros.append(r)

            ws_impuestos = wb_impuestos.create_sheet("CANTIDADES_ACTOS_GG_IMPUESTOS")
            for r in dataframe_to_rows(df_impuestos, index=False, header=True):
                ws_impuestos.append(r)

            wb_cobros.save(output_file_cobros)
            wb_impuestos.save(output_file_impuestos)

            messagebox.showinfo("Proceso completado", f"El archivo Excel ha sido actualizado exitosamente.\nArchivos filtrados por COBROS: {len(df_cobros)}\nArchivos filtrados por IMPUESTOS: {len(df_impuestos)}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al procesar la base de datos:\n{str(e)}")

    def abrir_directorio(self):
        output_directory = r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\REPORTES_ACTOS_DIARIOS_SDH_2024"
        subprocess.Popen(['explorer', output_directory])

    def center_window(self):
        self.update_idletasks()  # Actualiza la geometría de la ventana
        width = self.winfo_width()
        height = self.winfo_height()
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

# Instanciar y correr la ventana principal
if __name__ == "__main__":
    app = MainWindow()
    app.mainloop()
