import os
import tkinter as tk
from tkinter import Image, filedialog, messagebox, ttk
import io
from openpyxl import load_workbook
from PyPDF4 import PdfFileReader, PdfFileWriter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from PIL import Image, ImageTk
import threading
import time

# Lista para almacenar el estado compartido
estado_compartido = [False]

# Función para realizar un efecto de fade-in en la ventana
def fade_in(window):
    for alpha in range(0, 100):
        window.attributes('-alpha', alpha / 100)
        window.update_idletasks()
        time.sleep(0.01)

# Función para realizar un efecto de fade-out en la ventana
def fade_out(window):
    for alpha in range(100, 0, -1):
        window.attributes('-alpha', alpha / 100)
        window.update_idletasks()
        time.sleep(0.01)

def cargar_archivo_excel(lbl_archivo_excel):
    archivo_excel = filedialog.askopenfilename(title="Seleccione el archivo Excel", filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if archivo_excel:
        lbl_archivo_excel.delete(0, tk.END)
        lbl_archivo_excel.insert(0, archivo_excel)

def seleccionar_carpeta_salida(lbl_carpeta_salida):
    carpeta_salida = filedialog.askdirectory(title="Seleccione la carpeta de salida")
    if carpeta_salida:
        lbl_carpeta_salida.delete(0, tk.END)
        lbl_carpeta_salida.insert(0, carpeta_salida)

def limpiar_selecciones(lbl_archivo_excel, lbl_carpeta_salida):
    lbl_archivo_excel.delete(0, tk.END)
    lbl_carpeta_salida.delete(0, tk.END)

def cerrar_aplicacion(root):
    # Establece la bandera de estado compartido a True
    estado_compartido[0] = True
    # Realiza una transición de fade-out antes de cerrar la ventana principal
    fade_out(root)
    # Cierra la ventana principal después del fade-out
    root.destroy()

def mostrar_ventana_progreso(root):
    ventana_progreso = tk.Toplevel(root)
    ventana_progreso.title("Procesando archivos")
    # Inicia la ventana emergente con transparencia cero para efecto fade-in
    ventana_progreso.attributes('-alpha', 0)
    # Realiza una transición de fade-in para mostrar la ventana emergente gradualmente
    fade_in(ventana_progreso)
    
    # Ajusta el tamaño de la ventana emergente
    ventana_progreso.geometry("300x100")
    
    lbl_estado = tk.Label(ventana_progreso, text="Procesando archivos...")
    lbl_estado.pack()

    lbl_porcentaje = tk.Label(ventana_progreso, text="")
    lbl_porcentaje.pack()

    progress_bar = ttk.Progressbar(ventana_progreso, mode='determinate', length=260)
    progress_bar.pack()

    return ventana_progreso, lbl_estado, lbl_porcentaje, progress_bar

def procesar_archivos(root, lbl_archivo_excel, lbl_carpeta_salida):
    archivo_excel = lbl_archivo_excel.get()
    carpeta_salida = lbl_carpeta_salida.get()

    if archivo_excel and carpeta_salida:
        ventana_progreso, lbl_estado, lbl_porcentaje, progress_bar = mostrar_ventana_progreso(root)

        def process():
            # Accede al estado compartido usando la lista
            if estado_compartido[0]:
                messagebox.showinfo("Proceso interrumpido", "El proceso ha sido detenido.")
                ventana_progreso.destroy()
                return

            workbook = load_workbook(archivo_excel)
            worksheet = workbook["Sheet1"]
            max_row = worksheet.max_row

            for row in range(1, max_row + 1):
                pdf_name = worksheet.cell(row, 1).value
                pdf_path = worksheet.cell(row, 2).value
                sticker_name = worksheet.cell(row, 3).value
                sticker_path = worksheet.cell(row, 4).value

                if pdf_path is None:
                    messagebox.showwarning("Error", "La ruta del archivo PDF está vacía en la fila {}".format(row))
                    continue

                with open(pdf_path, 'rb') as pdf_file:
                    pdf_reader = PdfFileReader(pdf_file)
                    pdf_writer = PdfFileWriter()

                    is_first_page = True

                    for page_num in range(pdf_reader.getNumPages()):
                        page = pdf_reader.getPage(page_num)

                        if is_first_page:
                            packet = io.BytesIO()
                            can = canvas.Canvas(packet, pagesize=letter)
                            page_height_cm = float(page.mediaBox.getHeight()) / 28.35

                            if page_height_cm > 42:
                                can.drawImage(sticker_path, 6, 40.50*cm, 7.2*cm, 2.4*cm)
                            elif page_height_cm > 39:
                                can.drawImage(sticker_path, 6, 39.10*cm, 7.2*cm, 2.4*cm)
                            elif page_height_cm > 35:
                                can.drawImage(sticker_path, 6, 32.95*cm, 7.2*cm, 2.4*cm)
                            elif page_height_cm > 33:
                                can.drawImage(sticker_path, 6, 30.35*cm, 7.2*cm, 2.4*cm)
                            elif page_height_cm > 29:
                                can.drawImage(sticker_path, 6, 27.15*cm, 7.2*cm, 2.4*cm)
                            elif page_height_cm > 27:
                                can.drawImage(sticker_path, 6, 25.35*cm, 7.2*cm, 2.4*cm)
                            else:
                                can.drawImage(sticker_path, 6, 23.75*cm, 7.2*cm, 2.4*cm)
                            can.save()
                            packet.seek(0)
                            overlay = PdfFileReader(packet).getPage(0)
                            page.mergePage(overlay)
                            is_first_page = False

                        pdf_writer.addPage(page)

                    output_filename = os.path.join(carpeta_salida, pdf_name)
                    with open(output_filename, 'wb') as output:
                        pdf_writer.write(output)

                # Calcular el porcentaje de progreso
                porcentaje_progreso = (row / max_row) * 100
                # Actualizar la barra de progreso con el porcentaje de progreso
                progress_bar['value'] = porcentaje_progreso
                lbl_porcentaje.config(text="{}%".format(int(porcentaje_progreso)))
                ventana_progreso.update_idletasks()

            # Realiza una transición de fade-out antes de cerrar la ventana de progreso
            fade_out(ventana_progreso)
            ventana_progreso.destroy()
            messagebox.showinfo("Proceso terminado", "El proceso ha terminado satisfactoriamente.")

        threading.Thread(target=process).start()
    else:
        messagebox.showwarning("Error", "Por favor, seleccione el archivo Excel y la carpeta de salida antes de procesar.")


def main():
    root = tk.Tk()
    root.title("Integracion de Archivos")
    root.geometry("650x330")  # Establece el tamaño de la ventana a 650*330 píxeles
    # Calcula el ancho y alto de la pantalla
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    icon_image = Image.open(r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\ENTORNO_GUI\Background\INC.ico")
    icon_photo = ImageTk.PhotoImage(icon_image)
    root.iconphoto(False, icon_photo)

    # Calcula las coordenadas "x" e "y" para centrar la ventana en la pantalla
    x = (screen_width - 650) // 2  # Ancho de la ventana es 650
    y = (screen_height - 330) //2  # Ancho de la ventana es 330
    root.geometry("650x330+{}+{}".format(x, y))  # Establece el tamaño y la posición de la ventana

    # Frame para contener la imagen de fondo
    background_frame = tk.Frame(root)
    background_frame.place(relwidth=1, relheight=1)

    # Cargar y mostrar la imagen de fondo
    background_image = tk.PhotoImage(file=r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\ENTORNO_GUI\Background\fnd.png")
    background_label = tk.Label(background_frame, image=background_image)
    background_label.place(x=0, y=15, relwidth=1, relheight=1)

    # Frame para los elementos de entrada y botones
    input_frame = tk.Frame(root)
    input_frame.place(relx=0.5, rely=0.2, anchor='center')

    lbl_archivo_excel = tk.Entry(input_frame, width=60)
    lbl_archivo_excel.grid(row=0, column=0, padx=5, pady=5)

    btn_cargar_excel = tk.Button(input_frame, text="CARGAR FUENTE", command=lambda: cargar_archivo_excel(lbl_archivo_excel), bg="Gray", fg="white", font=("Arial", 11))
    btn_cargar_excel.grid(row=0, column=1, padx=5, pady=5)

    lbl_carpeta_salida = tk.Entry(input_frame, width=60)
    lbl_carpeta_salida.grid(row=1, column=0, padx=5, pady=5)

    btn_seleccionar_carpeta_salida = tk.Button(input_frame, text="CARPETA SALIDA", command=lambda: seleccionar_carpeta_salida(lbl_carpeta_salida), bg="Gray", fg="white", font=("Arial", 11))
    btn_seleccionar_carpeta_salida.grid(row=1, column=1, padx=5, pady=5)

    btn_procesar = tk.Button(root, text="PROCESAR ARCHIVOS", command=lambda: procesar_archivos(root, lbl_archivo_excel, lbl_carpeta_salida), bg="green", fg="white", font=("Arial", 11),width=20)
    btn_procesar.place(relx=0.5, rely=0.9, anchor='center')

    btn_limpiar_selecciones = tk.Button(root, text="LIMPIAR", command=lambda: limpiar_selecciones(lbl_archivo_excel, lbl_carpeta_salida), bg="orange", font=("Arial", 11),width=20)
    btn_limpiar_selecciones.place(relx=0.2, rely=0.9, anchor='center')

    btn_cerrar = tk.Button(root, text="CERRAR", command=lambda: cerrar_aplicacion(root), bg="red", fg="white", font=("Arial", 11,),width=20)
    btn_cerrar.place(relx=0.8, rely=0.9, anchor='center')

    # Inicia la aplicación con un efecto de fade-in gradual
    fade_in(root)

    root.mainloop()

if __name__ == "__main__":
    main()
