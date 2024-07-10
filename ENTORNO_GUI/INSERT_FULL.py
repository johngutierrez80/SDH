import os
import io
import tkinter as tk
from tkinter import Image, filedialog, messagebox, ttk
from openpyxl import load_workbook
from PyPDF4 import PdfFileReader, PdfFileWriter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from PIL import Image, ImageTk
import threading
import os

def abrir_directorio_original(lbl_carpeta_salida):
    carpeta_original = lbl_carpeta_salida.get()
    if carpeta_original:
        os.startfile(carpeta_original)
    else:
        messagebox.showwarning("Error", "Por favor, seleccione la carpeta original antes de abrir.")

def abrir_directorio_copia(lbl_carpeta_salida_copia):
    carpeta_copia = lbl_carpeta_salida_copia.get()
    if carpeta_copia:
        os.startfile(carpeta_copia)
    else:
        messagebox.showwarning("Error", "Por favor, seleccione la carpeta de copia antes de abrir.")


# Lista para almacenar el estado compartido entre hilos
estado_compartido = [False]

# Definir la variable thread_lock como global para controlar el acceso a recursos compartidos
thread_lock = threading.Lock()

# Función para hacer un elemento visible gradualmente con una animación de fade in
def fade_in(element, alpha=0.0):
    if alpha < 1.0:
        alpha += 0.05
        element.attributes("-alpha", alpha)
        element.after(50, lambda: fade_in(element, alpha))
    else:
        element.attributes("-alpha", 1.0)

# Función para hacer un elemento invisible gradualmente con una animación de fade out
def fade_out(element, alpha=1.0):
    if alpha > 0.0:
        alpha -= 0.05
        element.attributes("-alpha", alpha)
        element.after(50, lambda: fade_out(element, alpha))
    else:
        element.attributes("-alpha", 0.0)
        element.destroy()

# Función para cargar un archivo Excel y mostrar su ruta en una etiqueta
def cargar_archivo_excel(lbl_archivo_excel):
    archivo_excel = filedialog.askopenfilename(title="Seleccione el archivo Excel", filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if archivo_excel:  # Si se selecciona un archivo
        lbl_archivo_excel.delete(0, tk.END)  # Elimina el contenido existente en la etiqueta
        lbl_archivo_excel.insert(0, archivo_excel)  # Inserta la ruta del archivo seleccionado en la etiqueta

# Función para seleccionar una carpeta de salida y mostrar su ruta en una etiqueta
def seleccionar_carpeta_salida(lbl_carpeta_salida):
    carpeta_salida = filedialog.askdirectory(title="Seleccione la carpeta de salida")
    if carpeta_salida:  # Si se selecciona una carpeta
        lbl_carpeta_salida.delete(0, tk.END)  # Elimina el contenido existente en la etiqueta
        lbl_carpeta_salida.insert(0, carpeta_salida)  # Inserta la ruta de la carpeta seleccionada en la etiqueta

# Función para limpiar los campos de entrada
def limpiar_campos(*entries):
    for entry in entries:
        entry.delete(0, tk.END)  # Borra el contenido de cada campo de entrada

# Función para cerrar la aplicación y establecer la bandera de estado compartido en True
def cerrar_aplicacion(root):
    estado_compartido[0] = True  # Establece la bandera de estado compartido en True
    root.destroy()  # Cierra la ventana principal de la aplicación

# Función para procesar un archivo PDF superponiendo una etiqueta y guardando el resultado
def procesar_archivo(pdf_path, sticker_path, output_path):
    with open(pdf_path, 'rb') as input_file:  # Abre el archivo PDF en modo lectura binaria
        pdf_reader = PdfFileReader(input_file)  # Crea un objeto PdfFileReader para leer el archivo PDF
        pdf_writer = PdfFileWriter()  # Crea un objeto PdfFileWriter para escribir en un nuevo archivo PDF

        # Procesa solo la primera página del PDF
        page = pdf_reader.getPage(0)  # Obtiene la primera página del archivo PDF
        overlay_sticker(page, sticker_path)  # Superpone una etiqueta en la primera página
        pdf_writer.addPage(page)  # Añade la página procesada al nuevo archivo PDF

        for page_num in range(1, pdf_reader.getNumPages()):  # Itera sobre las páginas restantes del archivo PDF
            page = pdf_reader.getPage(page_num)  # Obtiene la página actual del archivo PDF
            pdf_writer.addPage(page)  # Añade la página sin modificar al nuevo archivo PDF

        with open(output_path, 'wb') as output_file:  # Abre el nuevo archivo PDF en modo escritura binaria
            pdf_writer.write(output_file)  # Escribe el contenido del objeto PdfFileWriter en el nuevo archivo PDF

# Función para superponer una etiqueta en una página PDF
def overlay_sticker(page, sticker_path):
    packet = io.BytesIO()  # Crea un flujo de bytes en memoria
    can = canvas.Canvas(packet, pagesize=letter)  # Crea un objeto canvas para dibujar en el flujo de bytes
    page_height_cm = float(page.mediaBox.getHeight()) / 28.35  # Obtiene la altura de la página en centímetros

    # Determina la posición de la etiqueta en función de la altura de la página
    if page_height_cm > 42:
        can.drawImage(sticker_path, 6, 40.50*cm, 7.2*cm, 2.4*cm)
    elif page_height_cm > 39:
        can.drawImage(sticker_path, 6, 39.10*cm, 7.2*cm, 2.4*cm)
    elif page_height_cm > 35:
        can.drawImage(sticker_path, 6, 32.95*cm, 7.2*cm, 2.4*cm)
    elif page_height_cm > 33:
        can.drawImage(sticker_path, 6, 30.35*cm, 7.2*cm, 2.4*cm)
    elif page_height_cm > 29:
        can.drawImage(sticker_path, 6, 27.15*cm, 7.2*cm, 2.4*cm)
    elif page_height_cm > 27:
        can.drawImage(sticker_path, 6, 25.35*cm, 7.2*cm, 2.4*cm)
    else:
        can.drawImage(sticker_path, 6, 23.75*cm, 7.2*cm, 2.4*cm)
    can.save()  # Guarda el contenido del objeto canvas en el flujo de bytes
    packet.seek(0)  # Coloca el cursor al inicio del flujo de bytes
    overlay = PdfFileReader(packet).getPage(0)  # Obtiene la página del flujo de bytes
    page.mergePage(overlay)  # Fusiona la página del PDF original con la página superpuesta

# Función para procesar archivos PDF en paralelo a partir de un archivo Excel
def procesar_archivos(archivo_excel, carpeta_salida, titulo, progress_bar, lbl_porcentaje, ventana_progreso, thread_lock):
    workbook = load_workbook(archivo_excel)  # Carga el archivo Excel
    worksheet = workbook["Sheet1"]  # Selecciona la hoja de trabajo del archivo Excel
    max_row = worksheet.max_row  # Obtiene el número máximo de filas en el archivo Excel

    for row in range(1, max_row + 1):  # Itera sobre todas las filas del archivo Excel
        pdf_name = worksheet.cell(row, 1).value  # Obtiene el nombre del PDF de la celda correspondiente
        pdf_path = worksheet.cell(row, 2).value  # Obtiene la ruta del PDF de la celda correspondiente
        sticker_name = worksheet.cell(row, 3).value  # Obtiene el nombre de la etiqueta de la celda correspondiente
        sticker_path = worksheet.cell(row, 4).value  # Obtiene la ruta de la etiqueta de la celda correspondiente

        if pdf_path is None:  # Si la ruta del PDF está vacía
            messagebox.showwarning("Error", "La ruta del archivo PDF está vacía en la fila {}".format(row))
            continue  # Continúa con la siguiente iteración del bucle

        output_filename = os.path.join(carpeta_salida, pdf_name)  # Genera la ruta de salida para el PDF procesado
        procesar_archivo(pdf_path, sticker_path, output_filename)  # Procesa el archivo PDF

        # Calcula y actualiza el porcentaje de progreso
        porcentaje_progreso = (row / max_row) * 100
        with thread_lock:  # Adquiere el bloqueo del hilo para evitar condiciones de carrera
            progress_bar['value'] = porcentaje_progreso  # Actualiza el valor de la barra de progreso
            lbl_porcentaje.config(text="{}%".format(int(porcentaje_progreso)))  # Actualiza el texto del porcentaje
            ventana_progreso.update_idletasks()  # Actualiza la interfaz gráfica de usuario

    ventana_progreso.destroy()  # Cierra la ventana de progreso una vez que se completa el procesamiento
    messagebox.showinfo("Proceso terminado", "El proceso ha terminado satisfactoriamente.")  # Muestra un mensaje de información al usuario

# Función para mostrar una ventana de progreso durante el procesamiento de archivos
def mostrar_ventana_progreso(root, titulo):
    ventana_progreso = tk.Toplevel(root)  # Crea una ventana emergente
    ventana_progreso.title(titulo)  # Establece el título de la ventana

    ventana_progreso.geometry("300x100")  # Establece las dimensiones de la ventana

    lbl_estado = tk.Label(ventana_progreso, text="Procesando archivos...")  # Crea una etiqueta para mostrar el estado
    lbl_estado.pack()  # Empaqueta la etiqueta en la ventana

    lbl_porcentaje = tk.Label(ventana_progreso, text="")  # Crea una etiqueta para mostrar el porcentaje de progreso
    lbl_porcentaje.pack()  # Empaqueta la etiqueta en la ventana

    progress_bar = ttk.Progressbar(ventana_progreso, mode='determinate', length=260)  # Crea una barra de progreso
    progress_bar.pack()  # Empaqueta la barra de progreso en la ventana

    return ventana_progreso, lbl_estado, lbl_porcentaje, progress_bar  # Devuelve los elementos de la ventana

# Función envoltorio para procesar archivos en un hilo separado y mostrar una ventana de progreso
def procesar_archivos_wrapper(archivo_excel, carpeta_salida, titulo):
    ventana_progreso, lbl_estado, lbl_porcentaje, progress_bar = mostrar_ventana_progreso(root, titulo)
    threading.Thread(target=procesar_archivos, args=(archivo_excel, carpeta_salida, titulo, progress_bar, lbl_porcentaje, ventana_progreso, thread_lock)).start()

# Función para procesar archivos originales al hacer clic en el botón correspondiente
def procesar_archivos_originales(root, lbl_archivo_excel, lbl_carpeta_salida):
    archivo_excel = lbl_archivo_excel.get()  # Obtiene la ruta del archivo Excel desde la etiqueta
    carpeta_salida = lbl_carpeta_salida.get()  # Obtiene la ruta de la carpeta de salida desde la etiqueta

    if archivo_excel and carpeta_salida:  # Si se han seleccionado tanto el archivo Excel como la carpeta de salida
        procesar_archivos_wrapper(archivo_excel, carpeta_salida, "Procesando archivos originales")  # Procesa los archivos en un hilo separado
    else:  # Si falta alguna de las selecciones
        messagebox.showwarning("Error", "Por favor, seleccione el archivo Excel y la carpeta de salida antes de procesar.")  # Muestra un mensaje de advertencia al usuario

# Función para procesar archivos de copia al hacer clic en el botón correspondiente
def procesar_archivos_copia(root, lbl_archivo_excel_copia, lbl_carpeta_salida_copia):
    archivo_excel_copia = lbl_archivo_excel_copia.get()  # Obtiene la ruta del archivo Excel de copia desde la etiqueta
    carpeta_salida_copia = lbl_carpeta_salida_copia.get()  # Obtiene la ruta de la carpeta de salida de copia desde la etiqueta

    if archivo_excel_copia and carpeta_salida_copia:  # Si se han seleccionado tanto el archivo Excel de copia como la carpeta de salida de copia
        procesar_archivos_wrapper(archivo_excel_copia, carpeta_salida_copia, "Procesando archivos copia")  # Procesa los archivos en un hilo separado
    else:  # Si falta alguna de las selecciones
        messagebox.showwarning("Error", "Por favor, seleccione el archivo Excel y la carpeta de salida de copia antes de procesar.")  # Muestra un mensaje de advertencia al usuario

# Función principal para ejecutar la aplicación
def main():
    global root  # Declara la variable root como global para que sea accesible desde otras funciones
    root = tk.Tk()  # Crea la ventana principal de la aplicación
    root.title("INTEGRACION DE LABELS")  # Establece el título de la ventana
    root.geometry("580x450")  # Establece el tamaño de la ventana

    # Animación de fade in para la ventana principal
    root.attributes("-alpha", 0.0)
    fade_in(root)

    # Función para centrar la ventana principal en el escritorio
    def centrar_ventana(root):
        root.update_idletasks()  # Actualiza la ventana para obtener sus dimensiones reales
        width = root.winfo_width()  # Obtiene el ancho de la ventana
        height = root.winfo_height()  # Obtiene la altura de la ventana
        x = (root.winfo_screenwidth() // 2) - (width // 2)  # Calcula la posición X para centrar la ventana
        y = (root.winfo_screenheight() // 2) - (height // 2)  # Calcula la posición Y para centrar la ventana
        root.geometry('{}x{}+{}+{}'.format(width, height, x, y))  # Establece la geometría centrada de la ventana
        icon_image = Image.open(r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\ENTORNO_GUI\Background\INC.ico")
        icon_photo = ImageTk.PhotoImage(icon_image)
        root.iconphoto(False, icon_photo)    
    # Carga la imagen de fondo
    img_path = r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\ENTORNO_GUI\Background\fnd2.png"
    img = tk.PhotoImage(file=img_path)

    # Crea un lienzo para colocar la imagen de fondo
    canvas = tk.Canvas(root, width=img.width(), height=img.height())
    canvas.pack(fill="both", expand=True)

    # Establece la imagen de fondo en el lienzo
    canvas.create_image(80, 20, anchor="nw", image=img)

    # Frame para los elementos de entrada y botones
    input_frame = tk.Frame(root)
    input_frame.pack(pady=20)

    # Crear campos de entrada y botones para los archivos originales y de copia
   # Crear campos de entrada y botones para los archivos originales y de copia
    lbl_archivo_excel = tk.Entry(input_frame, width=60)
    lbl_archivo_excel.grid(row=0, column=0, padx=5, pady=5)

    btn_cargar_excel = tk.Button(input_frame, text="FUENTE ORIGINAL", command=lambda: cargar_archivo_excel(lbl_archivo_excel), bg="Gray", fg="white", font=("Arial", 11), width=15)  # Establece el ancho a 15
    btn_cargar_excel.grid(row=0, column=1, padx=5, pady=5)

    lbl_carpeta_salida = tk.Entry(input_frame, width=60)
    lbl_carpeta_salida.grid(row=1, column=0, padx=5, pady=5)

    btn_seleccionar_carpeta_salida = tk.Button(input_frame, text="CARPETA ORIGINAL", command=lambda: seleccionar_carpeta_salida(lbl_carpeta_salida), bg="Gray", fg="white", font=("Arial", 11), width=15)  # Establece el ancho a 15
    btn_seleccionar_carpeta_salida.grid(row=1, column=1, padx=5, pady=5)

    lbl_archivo_excel_copia = tk.Entry(input_frame, width=60)
    lbl_archivo_excel_copia.grid(row=2, column=0, padx=5, pady=5)

    btn_cargar_excel_copia = tk.Button(input_frame, text="FUENTE COPIA", command=lambda: cargar_archivo_excel(lbl_archivo_excel_copia), bg="Gray", fg="white", font=("Arial", 11), width=15)  # Establece el ancho a 15
    btn_cargar_excel_copia.grid(row=2, column=1, padx=5, pady=5)

    lbl_carpeta_salida_copia = tk.Entry(input_frame, width=60)
    lbl_carpeta_salida_copia.grid(row=3, column=0, padx=5, pady=5)

    btn_seleccionar_carpeta_salida_copia = tk.Button(input_frame, text="CARPETA COPIA", command=lambda: seleccionar_carpeta_salida(lbl_carpeta_salida_copia), bg="Gray", fg="white", font=("Arial", 11), width=15)  # Establece el ancho a 15
    btn_seleccionar_carpeta_salida_copia.grid(row=3, column=1, padx=5, pady=5)


    # Botones para procesar archivos originales y de copia
    btn_procesar_originales = tk.Button(root, text="PROCESAR ORIGINALES", command=lambda: procesar_archivos_originales(root, lbl_archivo_excel, lbl_carpeta_salida), bg="green", fg="white", font=("Arial", 12),width=25)
    btn_procesar_originales.pack(pady=10)

    btn_procesar_copia = tk.Button(root, text="PROCESAR COPIA", command=lambda: procesar_archivos_copia(root, lbl_archivo_excel_copia, lbl_carpeta_salida_copia), bg="blue", fg="white", font=("Arial", 12),width=25)
    btn_procesar_copia.pack(pady=10)

    # Botón para limpiar las selecciones
    btn_limpiar_selecciones = tk.Button(root, text="LIMPIAR", command=lambda: limpiar_campos(lbl_archivo_excel, lbl_carpeta_salida, lbl_archivo_excel_copia, lbl_carpeta_salida_copia), bg="orange", font=("Arial", 11),width=15)
    btn_limpiar_selecciones.place(relx=0.85, rely=0.83, anchor='center')

    # Botón para cerrar la aplicación
    btn_cerrar = tk.Button(root, text="CERRAR", command=lambda: cerrar_aplicacion(root), bg="red", fg="white", font=("Arial", 11),width=15)
    btn_cerrar.place(relx=0.85, rely=0.93, anchor='center')

    btn_abrir_directorio_original = tk.Button(root, text="Abrir Original", command=lambda: abrir_directorio_original(lbl_carpeta_salida), bg="gray", fg="white", font=("Arial", 10),width=15)
    btn_abrir_directorio_original.place(relx=0.15, rely=0.82, anchor='center')

    btn_abrir_directorio_copia = tk.Button(root, text="Abrir Copia", command=lambda: abrir_directorio_copia(lbl_carpeta_salida_copia), bg="gray", fg="white", font=("Arial", 10),width=15)
    btn_abrir_directorio_copia.place(relx=0.15, rely=0.94, anchor='center')


    # Centrar la ventana principal
    centrar_ventana(root)

    root.mainloop()  # Ejecuta el bucle principal de la aplicación

# Verifica si el script se está ejecutando directamente y llama a la función main en ese caso
if __name__ == "__main__":
    main()  # Ejecuta la función main() para iniciar la aplicación.
