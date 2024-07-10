import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import fitz  
from math import ceil  
import tkinter as tk
from tkinter import Button, Label, Entry, messagebox, filedialog, ttk
import time
import subprocess
from openpyxl.styles import PatternFill
from PIL import Image, ImageTk

# Ruta predeterminada donde se encuentra el archivo Excel
DEFAULT_EXCEL_FOLDER = r'\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK'

# Definir colores
white_font = Font(color="FFFFFF")  # Color de fuente blanco
black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Fondo negro

# Función para realizar un efecto de fade-in en la ventana
def fade_in(window):
    for alpha in range(0, 100):
        window.attributes('-alpha', alpha / 100)
        window.update_idletasks()
        time.sleep(0.01)

# Función para visualizar archivo Num_pag
def abrir_archivo_excel():
    if archivo_generado:
        subprocess.Popen([archivo_generado], shell=True)
    else:
        messagebox.showerror("Error", "No se ha generado ningún archivo Excel.")

# Función para realizar un efecto de fade-out en la ventana
def fade_out(window):
    for alpha in range(100, 0, -1):
        window.attributes('-alpha', alpha / 100)
        window.update_idletasks()
        time.sleep(0.01)

def procesar_archivos():
    global archivo_generado  # Define la variable globalmente
    ubicacion_carpeta = carpeta_entry.get()
    archivo_excel = os.path.join(DEFAULT_EXCEL_FOLDER, "NUM_PAG.xlsx")  # Usar la ruta predeterminada para el archivo Excel
    salida_carpeta = salida_entry.get()

    if not ubicacion_carpeta:
        messagebox.showerror("Error", "Por favor, ingrese la ubicación de la carpeta.")
        return
    
    if not os.path.exists(ubicacion_carpeta):
        messagebox.showerror("Error", f"El directorio {ubicacion_carpeta} no existe.")
        return
    
    if not salida_carpeta:
        messagebox.showerror("Error", "Por favor, ingrese la ubicación del directorio de salida.")
        return

    if not os.path.exists(salida_carpeta):
        messagebox.showerror("Error", f"El directorio de salida {salida_carpeta} no existe.")
        return

    archivos_pdf = [archivo for archivo in os.listdir(ubicacion_carpeta) if archivo.endswith(".pdf")]

    wb = load_workbook(archivo_excel)
    ws = wb.active

    # Agregar encabezados a la primera fila del Excel
    ws.append(["NOMBRE ARCHIVO", "PAGINAS", "HOJAS"])

    # Eliminar las filas existentes desde la fila 2 en adelante
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    total_paginas = 0
    total_redondeo = 0

    # Calcula la posición de la ventana emergente en función de la posición de la ventana principal
    root.update_idletasks()
    x_root = root.winfo_x()
    y_root = root.winfo_y()

    progress_window = tk.Toplevel(root)
    progress_window.title("Progreso Resporte!...")

    # Animación fade-in para mostrar la ventana gradualmente
    fade_in(progress_window)

    progress_label_var = tk.StringVar()
    progress_label = Label(progress_window, textvariable=progress_label_var)
    progress_label.pack(padx=20, pady=10)

    progress_bar = ttk.Progressbar(progress_window, mode="determinate", maximum=len(archivos_pdf), length=260)
    progress_bar.pack(padx=20, pady=5, fill="x")

    # Calcula la posición de la ventana emergente
    width = progress_window.winfo_width()
    height = progress_window.winfo_height()
    x = x_root + (root.winfo_width() - width) // 2
    y = y_root + (root.winfo_height() - height) // 2
    progress_window.geometry(f"+{x}+{y}")

    for i, archivo_pdf in enumerate(archivos_pdf, start=1):
        archivo_completo = os.path.join(ubicacion_carpeta, archivo_pdf)
        nombre_archivo = os.path.basename(archivo_pdf)

        try:
            pdf = fitz.open(archivo_completo)  
            numero_paginas = pdf.page_count
            redondeo_maximo = ceil(numero_paginas / 2)  
            total_paginas += numero_paginas
            total_redondeo += redondeo_maximo
        except:
            numero_paginas = 'Error al obtener el número de páginas'
            redondeo_maximo = 'Error en el cálculo del redondeo máximo'

        siguiente_fila = get_next_empty_row(ws, 'A')
        ws['A' + str(siguiente_fila)] = nombre_archivo
        ws['B' + str(siguiente_fila)] = numero_paginas
        ws['C' + str(siguiente_fila)] = redondeo_maximo

        progress_label_var.set(f"Generando Reporte... {i}/{len(archivos_pdf)} ({(i/len(archivos_pdf))*100:.2f}%)")
        progress_bar["value"] = i
        progress_window.update()

    siguiente_fila = get_next_empty_row(ws, 'A')
    ws['A' + str(siguiente_fila)] = "TOTAL FINALES CLICKS/HOJAS"
    ws['B' + str(siguiente_fila)] = total_paginas
    ws['C' + str(siguiente_fila)] = total_redondeo

    # Escribir el texto y aplicar estilos
   
    red_font = Font(color="FF0000")
    ws['B' + str(siguiente_fila )].font = red_font
    ws['C' + str(siguiente_fila )].font = red_font

    # Guardar el archivo Excel en el directorio de salida seleccionado
    archivo_salida = os.path.join(salida_carpeta, "NUM_PAG_CARTAS.xlsx")
    wb.save(archivo_salida)
    archivo_generado = archivo_salida  # Almacena la ruta del archivo generado

    fade_out(progress_window)  # Realiza una transición de fade-out antes de cerrar la ventana
    progress_window.after(1000, progress_window.destroy)  # Espera un segundo antes de destruir la ventana
    messagebox.showinfo("Proceso Completado", f"El archivo Excel ha sido guardado exitosamente en {archivo_salida}.")

def get_next_empty_row(ws, column_letter):
    column_index = get_column_index(column_letter)
    for row in range(2, ws.max_row + 1):
        cell_value = ws[get_column_letter(column_index) + str(row)].value
        if cell_value is None:
            return row
    return ws.max_row + 1

def get_column_index(column_letter):
    return ord(column_letter.upper()) - 64

def carpeta_seleccionada():
    carpeta = filedialog.askdirectory()
    carpeta_entry.delete(0, tk.END)
    carpeta_entry.insert(0, carpeta)

def salida_seleccionada():
    carpeta = filedialog.askdirectory()
    salida_entry.delete(0, tk.END)
    salida_entry.insert(0, carpeta)

def limpiar_campos():
    carpeta_entry.delete(0, tk.END)
    salida_entry.delete(0, tk.END)

root = tk.Tk()
root.title("Reportes")

# Load background image
background_image = tk.PhotoImage(file=r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\ENTORNO_GUI\Background\Reportes.png")
background_label = tk.Label(root, image=background_image)
background_label.place(relwidth=1, relheight=1)
background_label.place(relx=0.5, rely=0.6, anchor=tk.CENTER)

# Calcula el ancho y el alto de la ventana
ancho_ventana = 640
alto_ventana = 420

# Obtiene el ancho y el alto de la pantalla
ancho_pantalla = root.winfo_screenwidth()
alto_pantalla = root.winfo_screenheight()

# Calcula la posición x y y para centrar la ventana
x = (ancho_pantalla // 2) - (ancho_ventana // 2)
y = (alto_pantalla // 2) - (alto_ventana // 2)

# Establece la geometría de la ventana para que aparezca centrada en la pantalla
root.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")
icon_image = Image.open(r"\\fjcaldas\SDH-Secretaria_Distrital_de_Hacienda\EJECUTABLES_PROCESOS_OK\ENTORNO_GUI\Background\INC.ico")
icon_photo = ImageTk.PhotoImage(icon_image)
root.iconphoto(False, icon_photo)

# Etiquetas y campos de entrada
carpeta_label = Label(root, text="Ubicación Renombrados:")
carpeta_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")
carpeta_entry = Entry(root, width=50)
carpeta_entry.grid(row=0, column=1, padx=5, pady=5)
carpeta_button = Button(root, text="Seleccionar Renombrados", command=carpeta_seleccionada, width=20)
carpeta_button.grid(row=0, column=2, padx=5, pady=5)

salida_label = Label(root, text="Directorio de Salida:")
salida_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
salida_entry = Entry(root, width=50)
salida_entry.grid(row=1, column=1, padx=5, pady=5)
salida_button = Button(root, text="Seleccionar Carpeta", command=salida_seleccionada, width=20)
salida_button.grid(row=1, column=2, padx=5, pady=5)

procesar_button = Button(root, text="Generar Reporte", command=procesar_archivos, bg="green", width=20)
procesar_button.grid(row=2, column=0, columnspan=3, pady=10)

limpiar_button = Button(root, text="Limpiar Selección", command=limpiar_campos, bg="orange", width=20)
limpiar_button.grid(row=3, column=0, pady=5, padx=5, sticky="we")

cerrar_button = Button(root, text="Cerrar", command=root.quit, bg="red", width=20)
cerrar_button.grid(row=3, column=2, pady=5, padx=5, sticky="we")

abrir_button = Button(root, text="Abrir Reporte", command=abrir_archivo_excel, bg="gray", width=20)
abrir_button.grid(row=3, column=1, pady=5, padx=5, sticky="we")

# Inicia la aplicación con un efecto de fade-in gradual
fade_in(root)

root.mainloop()
